# coding=gbk
# 主要用来放算法的地方
import baostock as bs
import datetime
import pymysql
import openpyxl
# def aaa():
#     db = pymysql.connect("47.111.24.112", "root", "withme_321", "test")
#     cursor = db.cursor()
#     get_K = get_Kline()
#     sql1="select  date,code from daban where price='0.00'"
#     cursor.execute(sql1)
#     end_list = cursor.fetchall()
#     for index in end_list:
#         try:
#             # print(index[0],index[1])
#             data_list=get_K.getKline(str(index[0]),str(index[0]),index[1])
#             price=data_list[0][5]
#             sql2="UPDATE  daban set price='%s' WHERE date='%s' and code='%s' "%(price,str(index[0]),index[1])
#             cursor.execute(sql2)
#             db.commit()
#         except:
#             continue
#
#     get_K.bs_close()
#     db.close()
# aaa()
# def bbb():
#     from pytdx.hq import TdxHq_API
#     from pytdx.params import TDXParams
#     api = TdxHq_API()
#     db = pymysql.connect("47.111.24.112", "root", "withme_321", "test")
#     cursor = db.cursor()
#     sql_date="SELECT DISTINCT date from bs where date>='%s' and date<='%s' ORDER BY date"%('2018-01-01','2021-05-11')
#     cursor.execute(sql_date)
#     num_days = cursor.fetchall()
#     with api.connect('119.147.212.81', 7709):
#         for day in num_days:
#             # print(day[0])
#             aa=api.get_history_minute_time_data(TDXParams.MARKET_SH, '000001',str(day[0]).replace('-',''))
#             morinig_price=aa[0]['price']
#             middle_price=aa[120]['price']
#             # print(morinig_price,middle_price)
#             if float(middle_price)>=float(morinig_price):type=1
#             else:type=0
#             tup_end = (day[0], morinig_price,middle_price, type)
#             sql = "INSERT INTO shzs VALUES ('%s','%s','%s','%s')" % tup_end
#             cursor.execute(sql)
#         db.commit()
#     db.close()
# bbb()
# def ccc():
#     import requests,time,json
#     db = pymysql.connect("47.111.24.112", "root", "withme_321", "test")
#     cursor = db.cursor()
#     s=requests.session()
#     date = datetime.datetime.strptime('2020-01-23', '%Y-%m-%d').date()
#     for i in range(-1,-100,-1):
#         num_day = datetime.timedelta(i)
#         oldday = date + num_day
#         timetemp=int(time.time() * 1000)
#         tup_canshu=str(oldday).replace("-","")
#         # print(tup_canshu)
#         url="http://push2ex.eastmoney.com/getTopicZTPool?cb=callbackdata4995459&ut=7eea3edcaed734bea9cbfc24409ed989&dpt=wz.ztzt&Pageindex=0&pagesize=60&sort=fbt%%3Aasc&date=%s&_=1623128546393"%tup_canshu
#         resp=s.get(url=url,headers={'Content-Type': 'application/json',
#                         'User-Agent': 'Mo Chrome/86.0.4240.75 Safari/537.36'})
#         content=resp.text.replace('callbackdata4995459(','').replace(');','')
#         if json.loads(content)['data']==None:continue
#         needdata=json.loads(content)['data']['pool']
#         for index in needdata:
#             name=index['n']
#             code=index['c']
#             price=float(index['p'])/1000
#             amount=float(index['amount'])
#             turn=str(index['hs'])[0:4]
#             updown=str(index['zdp'])[0:4]
#             shizhi,highdays,limittype,reason='0','null','null','null'
#             print(oldday,name,code,price,amount,turn)
#             data_tup = (oldday, code, name, price, updown, turn, shizhi, highdays, limittype, reason, amount)
#             sql = "INSERT INTO daban_zd VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % data_tup
#             cursor.execute(sql)
#         db.commit()
#     db.close()
# ccc()
class get_Kline(object):
    def __init__(self):
        self.lg = bs.login()
    def getKline(self,beginDate,endDate,code):
        if code[0] == '0':code = "sz." + code
        if code[0] == '6':code = 'sh.' + code
        if code[0] == '3':code = 'sz.' + code
        if code[0] == '2':code = 'sz.' + code
        if code[0] == '9':code = 'sh.' + code
        rs = bs.query_history_k_data_plus(code,
                                          "date,code,open,high,low,close,preclose,volume,amount,turn,pctChg",
                                          start_date=beginDate, end_date=endDate,
                                          frequency="d", adjustflag="2")
        data_list = []
        while (rs.error_code == '0') & rs.next():
            data_list.append(rs.get_row_data())
        # print(data_list)
        return data_list

    def testgetKline(self,beginDate, endDate, code,k):
        if code[0] == '0':code = "sz." + code
        if code[0] == '6':code = 'sh.' + code
        if code[0] == '3':code = 'sz.' + code
        if code[0] == '2':code = 'sz.' + code
        if code[0] == '9':code = 'sh.' + code
        rs = bs.query_history_k_data_plus(code,
                                          "date,time,code,open,high,low,close,volume,amount",
                                          start_date=beginDate, end_date=endDate,
                                          frequency=k, adjustflag="2")
        data_list = []
        while (rs.error_code == '0') & rs.next():
            # 获取一条记录，将记录合并在一起
            data_list.append(rs.get_row_data())
        return data_list
    def bs_close(self):
        bs.logout()

# def ddd():#下载所有股票代码
#     import requests,time,json
#     db = pymysql.connect("47.111.24.112", "root", "withme_321", "test")
#     cursor = db.cursor()
#     s=requests.session()
#     resp=s.get(url='http://api.k780.com/?app=finance.stock_list&category=hs&appkey=10003&sign=b59bc3ef6191eb9f747dd4e83c99f2a4&format=json',headers={'Content-Type': 'application/json',
#                         'User-Agent': 'Mo Chrome/86.0.4240.75 Safari/537.36'})
#     end=resp.json()['result']['lists']
#     for stock in end:
#         data_tup=(stock['symbol'][2:],stock['sname'])
#         # print(data_tup)
#         sql = "INSERT INTO stocks VALUES('%s','%s')" % data_tup
#         cursor.execute(sql)
#     db.commit()
#     db.close()
# ddd()
#下载历史深市中小板每日涨停股
# def eee():
#     get_K=get_Kline()
#     db = pymysql.connect("47.111.24.112", "root", "withme_321", "test")
#     cursor = db.cursor()
#     code_sql="SELECT * from stocks WHERE code like '00%'"
#     cursor.execute(code_sql)
#     code_list = cursor.fetchall()
#     date = datetime.datetime.strptime('2019-08-14', '%Y-%m-%d').date()
#     for i in range(-1,-200,-1):
#         num_day = datetime.timedelta(i)
#         oldday = str(date + num_day)
#         for stock in code_list:
#             code = stock[0]
#             name = stock[1]
#             if 'ST' in name:continue
#             if '退' in name:continue
#             data_list=get_K.getKline(oldday,oldday,code)
#             # print(data_list)
#             if data_list==[]:continue
#             price=float(data_list[0][5])
#             updown=float(data_list[0][10])
#             if data_list[0][9]=='':continue
#             turn=float(data_list[0][9])
#             amount=float(data_list[0][8])
#             if updown<9.95:continue
#             shizhi,highdays,limittype,reason='0','null','null','null'
#             data_tup = (oldday, code, name, price, updown, turn, shizhi, highdays, limittype, reason, amount)
#             print(data_tup)
#             sql = "INSERT INTO daban_zd VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % data_tup
#             cursor.execute(sql)
#         db.commit()
#     db.close()
# eee()
# def hhh():
#     get_K = get_Kline()
#     data_list1=get_K.getKline('2021-06-01','2021-06-07','002077')
#     # data_list2=get_K.testgetKline('2021-08-20','2021-08-20','003039','5')
#     print(data_list1)
#     # print(float(data_list2[29][3]))

def syl2_ruku():
    db = pymysql.connect("47.111.24.112", "root", "withme_321", "test")
    cursor = db.cursor()
    get_K = get_Kline()
    sql1="SELECT date,code from daban_zd WHERE updown >0 and code like'00%%' and date>='%s' and date<='%s'"%('2021-11-23','2021-12-08')
    cursor.execute(sql1)
    end_list = cursor.fetchall()
    for index in end_list:
        high_price, low_price, buy_price = [], [], 0
        code=index[1]
        day=index[0]
        data_list = get_K.getKline(str(day), str(day + datetime.timedelta(15)), code)
        # print(data_list)
        sell_day = data_list[2][0]  # 卖的日期
        data_list3 = get_K.testgetKline(sell_day, sell_day, code, '5')
        if data_list3 == []: continue

        for xxx in data_list3:
            high_price.append(float(xxx[4]))
            low_price.append(float(xxx[5]))
        high_price = high_price[3:47]
        low_price = low_price[3:29]
        highest_price, lowest_price = max(high_price), min(low_price)
        syl = (highest_price - float(data_list[1][5])) / float(data_list[1][5])

        if float(data_list[1][2]) < float(data_list[1][6]) and float(data_list[1][2]) * 0.992 > float(data_list[1][4]):
            buy_price = round(float(data_list[1][2]) * 0.992, 2)
        elif float(data_list[1][4]) <= round(float(data_list[1][6]) * 0.984, 2) < float(
                data_list[1][3]):  # 今日最低<上一日收盘价<今日最高
            buy_price = round(float(data_list[1][6]) * 0.984, 2)  # 买价，T日开盘价
        else:continue

        syl0 = (float(data_list[1][5]) - buy_price) / buy_price
        if syl >= 0.097:
            sell_price = float(data_list[1][5]) * 1.097
        elif syl0 < 0 and (float(data_list3[6][3]) - buy_price) / buy_price < -0.03:
            sell_price = float(data_list3[7][3])  # 1000与挂单价计算收益破-0.03，则1005卖
        else:
            if float(data_list3[29][3]) > buy_price:
                sell_price = float(data_list3[29][3])
            else:
                sell_price = float(data_list3[47][3])


        syl = "%.2f%% " % ((sell_price - buy_price) / buy_price * 100)  # 收益率
        print(syl,code,day)
        sql2="update daban_zd set syl2='%s' where date='%s'and code='%s'"%(syl,day,code)
        cursor.execute(sql2)
        db.commit()

    get_K.bs_close()
    db.close()
# syl2_ruku()



def suanfa1_use(cursor_handle,day,code):#给suanfa1需要用到的方法，用于计算第一个code无法买入时，更换第二个code
    sql="select code from daban_zd where date='%s'and model<>'None' ORDER BY dde_lookday desc limit 1"%day
    cursor_handle.execute(sql)
    code_tup = cursor_handle.fetchall()
    sencond_code=code_tup[0][0]
    if code==sencond_code:return None
    else:return sencond_code

def suanfa1_use2(day,code,get_K):#用于计算买价
    high_price, low_price,buy_price=[],[],0
    data_list = get_K.getKline(str(day), str(day + datetime.timedelta(15)), code)
    sell_day = data_list[2][0]
    data_list3 = get_K.testgetKline(sell_day, sell_day, code, '5')
    if data_list3 == []: return []
    if float(data_list[1][2]) < float(data_list[1][6])and float(data_list[1][2])*0.992>float(data_list[1][4]):
        buy_price = round(float(data_list[1][2])*0.992, 2)
    elif float(data_list[1][4]) <= round(float(data_list[1][6]) * 0.984, 2) < float(
            data_list[1][3]):  # 今日最低<上一日收盘价<今日最高
        buy_price = round(float(data_list[1][6]) * 0.984, 2)  # 买价，T日开盘价
    else:
        return None
    # 在[29][3]之前找出是否涨停
    for xxx in data_list3:
        high_price.append(float(xxx[4]))
        low_price.append(float(xxx[5]))
    high_price = high_price[3:47]
    high_price2=high_price[29:46]
    low_price = low_price[3:29]
    highest_price, lowest_price = max(high_price), min(low_price)
    syl = (highest_price - float(data_list[1][5])) / float(data_list[1][5])
    return syl,high_price2,low_price,data_list,data_list3,buy_price

def suanfa1():
    db = pymysql.connect("47.111.24.112", "root", "withme_321", "test")
    cursor = db.cursor()
    get_K = get_Kline()
    data={
'4/23/2021':'002240',
# '4/23/2021':'002630',
'4/26/2021':'000928',
'4/27/2021':'002612',
'4/28/2021':'002773',
'4/29/2021':'000718',
'4/30/2021':'000963',
'5/10/2021':'000630',
'5/11/2021':'000950',
'5/12/2021':'000625',
'5/13/2021':'002022',
'5/14/2021':'000625',
'5/17/2021':'002709',
# '5/17/2021': '002472',
'5/18/2021':'002997',
'5/19/2021':'002759',
'5/20/2021': '002235',
# '5/20/2021': '002387',
'5/21/2021':'000762',
'5/24/2021':'000402',
'5/25/2021':'000829',
'5/26/2021':'002585',
'5/27/2021':'002617',
'5/28/2021':'002284',
'5/31/2021':'002271',
# '5/31/2021':'002407',
'5/6/2021':'000155',
'5/7/2021':'002162',
'6/1/2021':'000504',
# '6/1/2021':'002647',
'6/10/2021':'000034',
# '6/10/2021': '002416',
'6/11/2021':'000909',
'6/15/2021':'002273',
'6/16/2021':'002383',
'6/17/2021': '003000',
'6/18/2021':'000158',
'6/2/2021':'003001',
'6/21/2021':'002683',
'6/22/2021':'002870',
'6/23/2021':'000625',
# '6/23/2021':'002812',
'6/24/2021':'000058',
'6/25/2021':'003022',
'6/28/2021':'002074',
# '6/28/2021': '002620',
'6/29/2021':'002759',
'6/3/2021':'000953',
'6/30/2021':'002326',
'6/4/2021':'002373',
'6/7/2021':'002156',
'6/8/2021':'002011',
# '6/8/2021':'002902',
'6/9/2021':'002625',
'7/1/2021':'000908',
# '7/1/2021':'002118',
'7/2/2021': '000034',
'7/12/2021':'000636',
'7/13/2021':'002411',
'7/14/2021':'000800',
'7/15/2021':'000977',
# '7/15/2021':'000519',
'7/16/2021':'000612',
'7/19/2021':'002396',
'7/20/2021':'002553',
'7/21/2021':'002108',
'7/22/2021':'002455',
'7/23/2021': '000507',
# '7/23/2021': '000572',
'7/26/2021':'002506',
'7/27/2021':'002463',
'7/28/2021':'002172',
# '7/28/2021':'000955',
'7/29/2021':'002129',
'7/30/2021':'000762',
'7/5/2021':'002460',
'7/6/2021':'002080',
'7/7/2021':'002074',
# '7/7/2021':'002108',
'7/8/2021':'000009',
# '7/8/2021':'002149',
'7/9/2021':'002709',
'8/9/2021': '002714',
'8/10/2021':'000576',
# '8/10/2021':'002416',
'8/11/2021':'002518',
'8/12/2021':'002312',
# '8/12/2021':'002326',
'8/13/2021':'000762',
'8/16/2021':'000572',
'8/17/2021':'000751',
'8/18/2021':'000738',
'8/19/2021':'002149',
'8/2/2021':'002594',
'8/20/2021':'000877',
'8/23/2021':'002465',
'8/24/2021':'000878',
'8/25/2021':'000821',
'8/26/2021':'002256',
'8/27/2021':'002497',
'8/3/2021':'002520',
'8/30/2021':'000762',
'8/31/2021':'002176',
'8/4/2021':'002594',
'8/5/2021':'002389',
# '8/5/2021':'002273',
'8/6/2021':'002407',
'9/1/2021':'000425',
# '9/1/2021': '000951',
'9/2/2021': '000301',
'9/3/2021': '002202',
'9/6/2021': '000825',
# '9/6/2021': '002320',
'9/7/2021': '002326',
'9/8/2021': '000966',
'9/9/2021': '002377',
'9/10/2021':'002240',
# '9/10/2021':'002411',
'9/13/2021':'002176',
'9/14/2021':'000009',
'9/15/2021':'002202',
'9/16/2021': '000912',
'9/17/2021':'000791',
'9/22/2021':'000883',
'9/23/2021':'000488',
'9/24/2021': '002459',
'9/27/2021': '000858',
'9/28/2021':'000591',
'9/29/2021':'002610',
'9/30/2021':'002709',
'10/8/2021':'000998',
'10/11/2021':'000836',
'10/12/2021':'000422',
'10/13/2021':'002326',
# '10/13/2021':'002434',
'10/14/2021': '002610',
'10/15/2021': '000423',
# '10/15/2021': '002105',
'10/18/2021': '002176',
'10/19/2021': '002261',
'10/20/2021': '002192',
'10/21/2021': '000422',
'10/22/2021': '002617',
'10/25/2021': '002074',
'10/26/2021': '000791',
'10/27/2021': '002202',
# '10/27/2021': '002090',
'10/28/2021': '002805',
'10/29/2021': '002312',
'11/01/2021': '002610',
'11/02/2021':'000982',
'11/03/2021': '002335',
# '11/03/2021': '002529',
'11/04/2021': '002202',
'11/05/2021': '002055',
'11/08/2021': '002326',
# '11/08/2021': '002245',
'11/09/2021': '002536',
'11/10/2021': '002254',
'11/11/2021':'002271',
'11/12/2021':'000049',
'11/15/2021': '002139',
'11/16/2021': '002263',
# '11/17/2021': '002466',
'11/17/2021': '002610',
'11/18/2021': '000155',
'11/22/2021': '002045',
'11/23/2021': '002056',
# '11/23/2021': '000982',
'11/24/2021': '003037',
'11/25/2021': '000400',
# '11/25/2021': '000969',
'11/26/2021': '002487',
'11/29/2021': '002056',
'11/30/2021': '002922',
'12/01/2021': '002006',
'12/02/2021': '002600',
'12/03/2021': '000591',
'12/06/2021': '002733',
'12/07/2021': '002708',
'12/08/2021': '002139',
'12/09/2021': '002084',
'12/10/2021': '002851',
'12/13/2021': '000951',
'12/14/2021': '002386',
'12/15/2021': '000795',
'12/16/2021': '000723',
          }
    data1={}
    money=50000
    sum = 0
    for k,v in data.items():
        klist=k.split('/')
        if len(klist[0])==1:klist[0]='0'+klist[0]
        if len(klist[1])==1:klist[1]='0'+klist[1]
        data1[klist[2]+'-'+klist[0]+'-'+klist[1]]=v
    for date,code in data1.items():
        sql2="select date,code from daban_zd where date='%s'and code='%s'"%(date,code)
        cursor.execute(sql2)
        stock_list= cursor.fetchall()
        for index in stock_list:
            day=index[0]
            code=index[1]

            returnend=suanfa1_use2(day,code,get_K)
            if returnend==[]:break
            elif returnend==None:break
                # sencond_code=suanfa1_use(cursor, str(day), code)
                # if sencond_code==None:break
                # returnend2=suanfa1_use2(day,sencond_code,get_K)
                # if returnend2 == [] or returnend2==None:break
                # else:syl,high_price,low_price,data_list,data_list3,buy_price=returnend2
            else:syl,high_price2,low_price,data_list,data_list3,buy_price=returnend
            # syl2=(float(data_list3[47][5])-float(data_list[2][6]))/float(data_list[2][6])
            # x = abs(float(dde_buyday) / float(dde_lookday))
            # if syl>=0.0992:
            #     sell_price=highest_price
            # elif float(dde_buyday)<0 and  0.8<x<1:
            # # !!!!满足dde条件后，卖日float(data_list3[3][3])是绿色的，则float(data_list3[3][3])价格卖，不是则使用上/下两个卖价
            #     sell_price=float(data_list3[3][3])
            #     syl2 = "%.2f%% " % ((sell_price - buy_price) / buy_price * 100)  # 收益率
            # else:sell_price=float(data_list3[47][3])

            syl0=(float(data_list[1][5]) - buy_price) / buy_price
            if syl>=0.097:
                sell_price=float(data_list[1][5])*1.097
            elif syl0<0 and (float(data_list3[6][3])-buy_price)/buy_price<-0.03:
                sell_price=float(data_list3[7][3])#1000与挂单价计算收益破-0.03，则1005卖
            else:
                if float(data_list3[29][3])>buy_price:#sell_price=float(data_list3[29][3])
                    if (float(data_list3[29][3])*1.025)<=float(max(high_price2)):
                        sell_price=float(data_list3[29][3])*1.025
                    else:sell_price=float(data_list3[47][3])
                else:sell_price = float(data_list3[47][3])
            money = money * (1 + (sell_price - buy_price) / buy_price) * 0.9985
            sum += 1
            print(money, sum, code, day, buy_price, sell_price)
    db.close()
suanfa1()

                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                        # 概念里，交易量最大的开盘跌了买，然后其他开盘跌的买（可以根据跌的幅度或者交易量），最后就是交易量最大的-1.5%
def suanfa0():#选择非首榜的试试
    # workbook = openpyxl.load_workbook('C:\\Users\Administrator\Desktop\daochu.xlsx')
    # sheet = workbook['Sheet1']
    # # 读取指定的单元格数据
    # data_dic={}
    # for index in range(1,134):
    #     cell1 = sheet.cell(row=index, column=1).value
    #     cell2 = sheet.cell(row=index, column=2).value
    #     data_dic[str(cell1.date())]=cell2
    # print(data_dic)
    # db = pymysql.connect("47.111.24.112", "root", "withme_321", "test")
    # cursor = db.cursor()
    get_K = get_Kline()
    money = 50000
    sum = 0
    data={'2021-04-23': '002630', '2021-04-26': '002120', '2021-04-27': '002612', '2021-04-28': '002932', '2021-04-29': '000718', '2021-04-30': '000718', '2021-05-06': '002762', '2021-05-07': '002721', '2021-05-10': '000762', '2021-05-11': '000893', '2021-05-12': '000625', '2021-05-13': '000158', '2021-05-14': '000625', '2021-05-17': '003027', '2021-05-18': '000829', '2021-05-19': '000066', '2021-05-20': '002248', '2021-05-21': '002227', '2021-05-24': '002471', '2021-05-25': '002717', '2021-05-26': '000762', '2021-05-27': '000966', '2021-05-28': '002581', '2021-05-31': '002617', '2021-06-01': '002957', '2021-06-02': '000415', '2021-06-03': '003031', '2021-06-04': '002178', '2021-06-07': '002326', '2021-06-08': '002902', '2021-06-09': '002136', '2021-06-10': '000009', '2021-06-11': '002862', '2021-06-15': '002530', '2021-06-17': '002993', '2021-06-18': '000158', '2021-06-21': '002023', '2021-06-22': '000848', '2021-06-24': '000301', '2021-06-25': '002482', '2021-06-28': '002803', '2021-06-29': '002119', '2021-06-30': '002892', '2021-07-01': '000062', '2021-07-05': '002176', '2021-07-06': '000935', '2021-07-07': '002108', '2021-07-08': '003022', '2021-07-09': '002472', '2021-07-12': '000902', '2021-07-13': '002411', '2021-07-14': '002732', '2021-07-15': '002892', '2021-07-16': '002286', '2021-07-19': '000935', '2021-07-20': '002580', '2021-07-21': '000762', '2021-07-22': '003038', '2021-07-23': '002224', '2021-07-26': '002170', '2021-07-27': '002201', '2021-07-28': '002580', '2021-07-29': '002435', '2021-07-30': '002497', '2021-08-02': '002335', '2021-08-03': '002520', '2021-08-04': '002594', '2021-08-05': '002389', '2021-08-06': '002819', '2021-08-09': '002006', '2021-08-10': '000420', '2021-08-11': '000723', '2021-08-12': '000676', '2021-08-13': '000676', '2021-08-16': '000829', '2021-08-18': '000762', '2021-08-19': '003040', '2021-08-20': '002939', '2021-08-23': '002015', '2021-08-24': '000629', '2021-08-25': '000821', '2021-08-26': '002256', '2021-08-27': '002598', '2021-08-30': '002631', '2021-08-31': '002240', '2021-09-01': '002208', '2021-09-02': '000983', '2021-09-03': '002202', '2021-09-06': '002353', '2021-09-07': '002079', '2021-09-08': '002597', '2021-09-09': '000825', '2021-09-10': '002155', '2021-09-13': '000807', '2021-09-14': '002529', '2021-09-15': '001896', '2021-09-16': '002606', '2021-09-17': '000912', '2021-09-22': '000655', '2021-09-23': '002153', '2021-09-27': '002366', '2021-09-28': '000993', '2021-09-29': '002865', '2021-09-30': '000767', '2021-10-08': '000552', '2021-10-11': '002505', '2021-10-12': '000899', '2021-10-13': '000150', '2021-10-14': '002607', '2021-10-18': '000301', '2021-10-19': '002935', '2021-10-20': '000829', '2021-10-21': '000422', '2021-10-22': '002655', '2021-10-25': '002665', '2021-10-26': '002655', '2021-10-28': '002805', '2021-10-29': '002283', '2021-11-01': '002837', '2021-11-02': '002916', '2021-11-03': '002529', '2021-11-04': '002454', '2021-11-05': '002540', '2021-11-08': '002045', '2021-11-09': '002073', '2021-11-10': '002254', '2021-11-11': '002017', '2021-11-12': '002073', '2021-11-15': '002405', '2021-11-16': '002728', '2021-11-17': '002056', '2021-11-18': '002131',
          '2021-11-22': '002664'}
    for day,code in data.items():
        # print(datetime.date.strptime(day, "%Y-%m-%d").date())
        returnend = suanfa1_use2(datetime.datetime.strptime(day, "%Y-%m-%d").date(), code, get_K)
        if returnend == []:
            continue
        elif returnend == None:
            continue
        # sencond_code=suanfa1_use(cursor, str(day), code)
        # if sencond_code==None:break
        # returnend2=suanfa1_use2(day,sencond_code,get_K)
        # if returnend2 == [] or returnend2==None:break
        # else:syl,high_price,low_price,data_list,data_list3,buy_price=returnend2
        else:
            syl, high_price, low_price, data_list, data_list3, buy_price = returnend

        syl0 = (float(data_list[1][5]) - buy_price) / buy_price
        if syl >= 0.097:
            sell_price = float(data_list[1][5]) * 1.097
        elif syl0 < 0 and (float(data_list3[6][3]) - buy_price) / buy_price < -0.03:
            sell_price = float(data_list3[7][3])  # 1000与挂单价计算收益破-0.03，则1005卖
        else:
            if float(data_list3[29][3]) > buy_price:
                sell_price = float(data_list3[29][3])
            else:
                sell_price = float(data_list3[47][3])

        money = money * (1 + (sell_price - buy_price) / buy_price) * 0.9985
        sum += 1
        print(money, sum, code, day, buy_price, sell_price)
# suanfa0()

def suanfa2():
    money,sum=50000,0
    get_K = get_Kline()
    data={'2021-09-29': ['002610', 59700, -90800],'2021-09-30': ['002709',53700, -44800],'2021-10-08': ['000998',17600,-20400],
          '2021-10-11': ['000836', 5334.09, -3152.31],'2021-10-12': ['000422',31900,734.84],'2021-10-13': ['002326',21800,-2480.54],
          '2021-10-14': ['002610', 41300, -15400],'2021-10-15': ['000423',17800,-23900],'2021-10-18': ['002610',56800,-1395.54],
          '2021-10-19': ['002261', 15000, -14900],'2021-10-20': ['002192',23000,-10000],'2021-10-21': ['000422',38000.00,-68100],
          '2021-10-22': ['002617', 26900, -684.26],'2021-10-25': ['000009', 81200, -4641.81],
        '2021-07-27': ['002902', 6773.72, 2340.1], '2021-08-09': ['002149', 20400, -24300], '2021-06-15': ['002273', 38700, -50100], '2021-06-03': ['000953', 801.04, -539.45], '2021-07-08': ['000009', 15800, -32000], '2021-07-06': ['002080', 14600, -7745.91], '2021-08-05': ['002074', 71000, -60700], '2021-08-27': ['002497', 65100, -56100], '2021-08-20': ['000155', 57200, -35600], '2021-09-10': ['002176', 104300, 22200], '2021-09-24': ['000862', 23400, 23400], '2021-05-14': ['000625', 118400, -120000], '2021-09-23': ['000488', 28000, -18300], '2021-06-25': ['002218', 4673.24, 4707.9], '2021-09-14': ['000009', 73100, -41700], '2021-08-24': ['000878', 26100, -16400], '2021-05-12': ['000625', 114800, -81900], '2021-07-30': ['000762', 39900, -52600], '2021-07-13': ['002411', 13800, -13800], '2021-09-02': ['002202', 64900, 97900], '2021-04-27': ['002612', 16600, -6781.68], '2021-08-04': ['002594', 232500, -97600], '2021-07-23': ['002623', 8411.46, -3089.84], '2021-06-16': ['002383', 7872.95, -14100], '2021-08-23': ['002218', 13300, -8950.47], '2021-06-08': ['002011', 3836.44, -3010.12], '2021-07-28': ['002172', 9658.95, -963.07], '2021-09-13': ['002312', 23300, -36600], '2021-08-19': ['002149', 21400, 5020.89], '2021-04-30': ['000963', 68300, -160800], '2021-06-11': ['000909', 4795.77, -3663.04], '2021-07-22': ['002407', 28400, -43500], '2021-05-11': ['000950', 7371.6, -6074.29], '2021-08-12': ['000792', 32700, -38300], '2021-04-29': ['000718', 15000, 3317.28], '2021-05-21': ['000762', 26100, -29800], '2021-06-09': ['002625', 19400, -959.28], '2021-09-16': ['000731', 16300, 4640.84], '2021-06-07': ['002156', 39900, -27600], '2021-08-17': ['001207', 11000, 4877.84], '2021-06-10': ['000034', 22900, -8850.74], '2021-07-01': ['000908', 12100, -9715.09], '2021-07-26': ['002506', 39200, -27200], '2021-06-30': ['002326', 22300, -20700], '2021-06-28': ['002218', 4707.9, 415.28], '2021-08-13': ['000762', 135700, -25200], '2021-05-24': ['000402', 10800, -3358.6], '2021-05-25': ['000793', 8648.86, -4588.6], '2021-06-24': ['002276', 9355.61, -6290.47], '2021-06-22': ['002623', 4375.28, -328.32], '2021-08-18': ['000738', 26200, -1711.03], '2021-07-09': ['002340', 183900, -97100], '2021-08-10': ['002625', 25200, -19200], '2021-05-27': ['002617', 23000, -3568.71], '2021-05-06': ['000825', 21500, -15500], '2021-09-01': ['000425', 45400, -40300], '2021-05-17': ['002709', 17000, -2765.7], '2021-06-29': ['002759', 13300, -13400], '2021-05-28': ['002407', 43000, 8206.39], '2021-06-01': ['002319', 1312.63, 1397.64], '2021-04-28': ['002773', 11700, -11300], '2021-05-18': ['002997', 7936.82, -5152.66], '2021-07-12': ['000636', 36000, -6788.8], '2021-07-16': ['000768', 59900, -7121.68], '2021-07-21': ['002466', 146600, 20800], '2021-07-20': ['000155', 60000, -37100], '2021-04-23': ['002340', 98900, -56100], '2021-07-02': ['002176', 19800, 6618.79], '2021-05-20': ['002235', 6354.89, -3988.65], '2021-06-23': ['000625', 104100, -51200], '2021-05-13': ['002022', 13300, -4077.56], '2021-08-11': ['002326', 30400, 7893.9], '2021-05-19': ['002759', 7906.08, -2591.69], '2021-08-31': ['000629', 25200, -23900], '2021-08-02': ['002594', 246100, -153800], '2021-06-02': ['003001', 8043.59, -6439.51], '2021-08-30': ['000762', 95500, -4522.85], '2021-08-16': ['000821', 13500, -11800], '2021-07-15': ['000825', 27500, 443.4], '2021-07-14': ['002581', 12200, -12800], '2021-07-07': ['002074', 33000, -9923.4], '2021-09-17': ['000791', 23100, 4527.61], '2021-06-18': ['000158', 38800, -65400], '2021-06-04': ['002373', 26500, -18000], '2021-08-25': ['000821', 15100, -4286.06], '2021-06-21': ['002683', 5611.18, 1031], '2021-05-10': ['000630', 56600, -42000], '2021-07-05': ['002460', 209100, -27800], '2021-08-06': ['002407', 97600, -87000], '2021-04-26': ['000908', 12400, 1497.38], '2021-05-07': ['002162', 25200, -19100], '2021-05-26': ['003039', 28400, 4435.84], '2021-06-17': ['002594', 138700, 90400], '2021-07-29': ['002466', 98300, 19400], '2021-09-22': ['000883', 21200, -9468.6], '2021-09-28': ['000591', 57100, -117500], '2021-09-27': ['000858', 354700, -35800], '2021-05-31': ['002585', 10200, 156.51], '2021-09-15': ['002202', 80000, -62000], '2021-08-03': ['002520', 11100, -4569.12], '2021-07-19': ['002421', 14300, -3794.25], '2021-08-26': ['000960', 19700, -14500]}
    for k,v in data.items():
        code=v[0]
        day = datetime.datetime.strptime(k, "%Y-%m-%d").date()
        data_list = get_K.getKline(str(day), str(day + datetime.timedelta(15)), code)

        sell_day = data_list[2][0]  # 卖的日期
        # 5分钟为单位的K线
        data_list3 = get_K.testgetKline(sell_day, sell_day, code, '5')
        if data_list3 == []: continue
        high_price = []
        data_list4 = get_K.testgetKline(sell_day, sell_day, code, '30')
        for xxx in data_list4:
            high_price.append(float(xxx[4]))
        high_price = high_price[0:4]
        if float(data_list[1][2]) < float(data_list[1][6]):buy_price = float(data_list[1][2])
        elif float(data_list[1][4])<=float(data_list[1][6])*0.984<float(data_list[1][3]):buy_price=float(data_list[1][6])*0.984 # 买价，T日开盘价
        else:continue
        highest_price = max(high_price)
        syl = (highest_price - float(data_list[1][5])) / float(data_list[1][5])
        if syl >= 0.0992:
            sell_price = highest_price
        # elif v[2]<0 and v[1]*0.7<=abs(v[2]):
        #     if buy_price*1.06<=highest_price:sell_price=buy_price*1.06
        #     else:sell_price = float(data_list3[29][3])
            # sell_price=float(data_list3[3][3])
        else:sell_price=float(data_list3[29][3])
        money = money * (1 + (sell_price - buy_price) / buy_price) * 0.9985
        sum += 1
        print(money, sum, code, day, buy_price, sell_price)
# suanfa2()

#尝试用龙虎榜最大交易额,失败的算法
def suanfa3():
    workbook = openpyxl.load_workbook('C:\\Users\Administrator\Desktop\\aa.xlsx')
    sheet = workbook['Sheet1']
    # 读取指定的单元格数据
    data_dic={}
    for index in range(1,113):
        cell1 = sheet.cell(row=index, column=1).value
        cell3 = sheet.cell(row=index, column=3).value
        # print(cell1.date())
        data_dic[cell1]=cell3
    money,sum=50000,0
    get_K = get_Kline()
    for k,v in data_dic.items():
        day = k.date()
        code=v
        data_list = get_K.getKline(str(day), str(day + datetime.timedelta(15)), code)

        sell_day = data_list[2][0]  # 卖的日期
        # 5分钟为单位的K线
        data_list3 = get_K.testgetKline(sell_day, sell_day, code, '5')
        if data_list3 == []: continue
        high_price = []
        data_list4 = get_K.testgetKline(sell_day, sell_day, code, '30')
        for xxx in data_list4:
            high_price.append(float(xxx[4]))
        high_price = high_price[0:4]
        if float(data_list[1][2]) < float(data_list[1][6]):buy_price = float(data_list[1][2])
        elif float(data_list[1][4]) <= float(data_list[1][6]) * 0.984< float(data_list[1][3]):buy_price = float(data_list[1][6]) * 0.984 # 买价，T日开盘价
        else:continue
        highest_price = max(high_price)
        syl = (highest_price - float(data_list[1][5])) / float(data_list[1][5])
        if syl >= 0.095:
            print(syl)
            sell_price = highest_price
        else:sell_price=float(data_list3[29][3])
        money = money * (1 + (sell_price - buy_price) / buy_price) * 0.9985
        sum += 1
        print(money, sum, code, day, buy_price, sell_price)
# suanfa3()
